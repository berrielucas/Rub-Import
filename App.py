import customtkinter as ctk
import requests as rqt
import openpyxl
from tkinter import *
from tkinter import filedialog
from time import sleep
import webbrowser
import threading
from queue import Queue
from CTkMessagebox import CTkMessagebox
import logging
from PIL import Image

ctk.set_appearance_mode("system")

logging.basicConfig(level=logging.INFO, filename="rubimport.log", format="%(asctime)s - %(levelname)s - %(message)s")

def position(win, width, height):
    largura_tela = win.winfo_screenwidth()
    altura_tela =  win.winfo_screenheight()
    posix = largura_tela//2 - width//2
    posiy = altura_tela//2 - height//2
    win.geometry(f"{width}x{height}+{posix}+{posiy}")


class App(ctk.CTk):
    def __init__(self, dicionario, dicionarioDois, link, origem, token):
        super().__init__()
        dados = {'link':f'{link}','token':f'{token}','origem':origem}
        logging.info(f'Conexão estabelecida - dados: {dados}')
        
        # Variáveis do sistema
        self.link = link
        self.origem = int(origem)
        self.token = token
        self.OfertasCursos = dicionarioDois
        self.dicionario = dicionario
        self.nomeModStr = ctk.StringVar(self)
        self.nomeUnidadeStr = ctk.StringVar(self)
        self.nomeNivelStr = ctk.StringVar(self)
        self.nomeTipoStr = ctk.StringVar(self)
        self.idUnidade = 0
        self.conectar = True
        self.oculto = True
        self.trancado = True

        img_lock = ctk.CTkImage(light_image=Image.open("./assets/lock-solid.png"), dark_image=Image.open("./assets/lock-solid-dark.png"), size=(15,16))
        img_unlock = ctk.CTkImage(light_image=Image.open("./assets/unlock-keyhole-solid-light.png"), dark_image=Image.open("./assets/unlock-keyhole-solid-dark.png"), size=(15,16))

        img_olho = ctk.CTkImage(light_image=Image.open("./assets/eye-solid-light.png"), dark_image=Image.open("./assets/eye-solid-dark.png"), size=(18,13))
        img_olho_oculto = ctk.CTkImage(light_image=Image.open("./assets/eye-slash-solid-light.png"), dark_image=Image.open("./assets/eye-slash-solid-dark.png"), size=(18,13))


        if self._get_appearance_mode() == 'light':
            self.fg_selects = '#D3D3D3'
            self.color_unselected = '#D3D3D3'
            self.color_terminal = '#363636'
            self.color_selected = '#A9A9A9'
            self.text_color = 'black'
            self.hover_butom = '#C0C0C0'
            self.border_color = 'black'
            self.btn_color = 'black'
            self.borderInt = 0
            ctk.set_appearance_mode('light')
        else:
            self.fg_selects = '#1C1C1C'
            self.color_unselected = '#363636'
            self.color_terminal = 'black'
            self.color_selected = '#1C1C1C'
            self.text_color = '#ffffff'
            self.hover_butom = '#4F4F4F'
            self.border_color = 'silver'
            self.btn_color = 'blue'
            self.borderInt = 0
            ctk.set_appearance_mode('dark')

        
        # Configuração da tela
        self.title("Rub Import")
        position(win=self, width=1000, height=500)
        self.grid_columnconfigure((2), weight=0)
        self.grid_columnconfigure((1,3), weight=1)
        self.grid_rowconfigure((1, 2), weight=1)
        self.iconbitmap('./assets/icon4.ico')


         # Funções do Sistema
        def atualizarDados(origem, token, link):

            params = {
                "origem":origem,
                "token":f"{token}"
            }

            response = rqt.request('get', url=f"{link}api/Curso/listarCursos", json=params, timeout=2)
            response.encoding
            response.text
            listaCurso = response.json()
            sleep(0.5)
            response = rqt.request('get', url=f"{link}api/Curso/listarOfertas", json=params, timeout=2)
            response.encoding
            response.text
            listaOferta = response.json()
            sleep(0.5)
            if listaOferta['success'] == False or listaCurso['success'] == False:
                logging.error('Erro ao atualizar dados')
                self.conectar = False
            else:
                self.OfertasCursos = {
                    'curso': listaCurso['dados'],
                    'oferta': listaOferta['dados']
                }
                listaCodigos = []
                for codigo in self.OfertasCursos['curso']:
                    listaCodigos.append(codigo['codigo'])
                self.codigosCurso = listaCodigos
                logging.info('Dados atualizados com sucesso')
                self.conectar = True

        def abrirArquivo():
            self.arquivo.set('Selecione o Arquivo')
            arquivo = filedialog.askopenfilename(title='Selecione o arquivo', filetypes=(('ecxel files', '*.xlsx'), ('All files', '*.*')))
            if arquivo != "":
                self.arquivo.set(arquivo)
        
        def escolherModalidade(value):
            if value != 'Selecione a Modalidade':
                self.nomeModStr.set(value)
            else:
                self.nomeModStr.set('')
    
        def escolherUnidade(value):
            if value != 'Selecione a Unidade':
                self.nomeUnidadeStr.set(value)
            else:
                self.nomeModStr.set('')
        
        def escolherNivel(value):
            if value != 'Selecione o Nível':
                self.nomeNivelStr.set(value)
            else:
                self.nomeModStr.set('')

        def escolherImportacao(value):
            self.nomeTipoStr.set(value)
        
        def abrirUrl():
            webbrowser.open('https://github.com/berrielucas/Rub-Import/blob/main/README.md')
        
        def acionarImportCurso():
            cadastrarCurso(file=self.arquivo_input.get(), link=self.link, origem=self.origem, token=self.token, tipo=self.switch_cod.get(), dic=self.OfertasCursos['curso'], masterTerminal=self.terminal, master=self, dic_cod_curso=self.codigosCurso)

            atualizarDados(link=self.link, origem=self.origem, token=self.token)
            self.update()

        def acionarImport():
            if self.nomeTipoStr.get() == 'Curso':
                acionarImportCurso()
            elif self.nomeTipoStr.get() == 'Oferta':
                acionarImportOferta()

        def acionarImportOferta():
            listaCodigosLoc = []
            for i in self.switchLoc_oferta.get():
                for local in self.dicionario['localOferta']:
                    if i == local['titulo']:
                        listaCodigosLoc.append(local['codigo'])
            
            for unidade in self.dicionario['unidade']:
                if self.nomeUnidadeStr.get() == unidade['nome']:
                    codUnidade = unidade['codigo']
                    self.idUnidade = unidade['id']

            for mod in self.dicionario['modalidade']:
                if self.nomeModStr.get() == mod['titulo']:
                    codModalidade = mod['codigo']
                    idModalidade = int(mod['id'])

            for nivel in self.dicionario['nivelEnsino']:
                if self.nomeNivelStr.get() == nivel['titulo']:
                    idNivel = int(nivel['id'])

            cadastrarOferta(file=self.arquivo_input.get(), link=self.link, origem=self.origem, token=self.token, codModalidade=codModalidade, idModalidade=idModalidade, codUnidade=codUnidade, listLocal=listaCodigosLoc, idNivel=idNivel, tipo=self.switch_cod.get(), dic=self.OfertasCursos['oferta'], masterTerminal=self.terminal, master=self, dic_curso=self.codigosCurso)

            atualizarDados(link=self.link, origem=self.origem, token=self.token)
            self.update()

        def desocultar():
            if self.oculto == True:
                self.token_input.configure(show="")
                self.btn_token.configure(image=img_olho)
                self.oculto = False
            else:
                self.token_input.configure(show="•")
                self.btn_token.configure(image=img_olho_oculto)
                self.oculto = True

        def destrancar():
            if self.trancado == True:
                self.token_input.configure(state='normal')
                self.origem_input.configure(state='normal')
                self.btn_tranca.configure(image=img_unlock)
                self.update()
                self.trancado = False
                self.editar_btn.grid(row=4, column=1, pady=(10,0), padx=(10), sticky="e", columnspan=2)
                self.update()
            else:
                self.token_input.configure(state='disable')
                self.origem_input.configure(state='disable')
                self.btn_tranca.configure(image=img_lock)
                self.update()
                self.trancado = True
                self.editar_btn.configure(fg_color=self.btn_color, state='normal')
                self.editar_btn.grid_forget()
                self.tokenStr.set(value=self.token)
                self.origemStr.set(value=self.origem)
                self.update()
                
        def alterarChave():
            dados = {'link':f'{self.link_input.get()}','token':f'{self.token_input.get()}','origem':self.origem_input.get()}
                    
            self.editar_btn.configure(state='disable', fg_color='#191970')
            self.token_input.configure(state='disable')
            self.origem_input.configure(state='disable')
            self.update()
            self.barra_editar.grid(column=0, row=5, columnspan=2, sticky='sew', pady=0, padx=15)
            self.update()

            threading.Thread(target=lambda:atualizarDados(link=self.link_input.get(), token=self.token_input.get(), origem=int(self.origem_input.get()))).start()

            self.after(10000, lambda:verifica(dados))
        
        def verifica(dados):
            if self.conectar == True:
                self.token = self.token_input.get()
                self.origem = self.origem_input.get()
                self.barra_editar.grid_forget()
                self.update()
                logging.info(f'Token/Origem atualizados com sucesso. Dados:{dados}')
                destrancar()
            elif self.conectar == False:
                self.barra_editar.grid_forget()
                self.update()
                self.btn_tranca.configure(image=img_unlock)
                self.token_input.configure(state='normal')
                self.origem_input.configure(state='normal')
                self.editar_btn.configure(state='normal', fg_color=self.btn_color)
                self.update()
                logging.error(f"Token/Origem sem permissão para se conectar. Dados:{dados}")
                self.tokenStr.set(value=self.token)
                self.origemStr.set(value=self.origem)
                alerta_permissao()


        # Atualzação de Listas
        self.nomeUnidade = ['Selecione a Unidade']
        self.nomeMod = ['Selecione a Modalidade']
        self.nomeNivel = ['Selecione o Nível']
        self.nomeLocal = []
        self.codigosCurso = []

        for codigo in self.OfertasCursos['curso']:
            self.codigosCurso.append(codigo['codigo'])

        for unidade in self.dicionario['unidade']:
            self.nomeUnidade.append(unidade['nome'])

        for mod in self.dicionario['modalidade']:
            self.nomeMod.append(mod['titulo'])

        for local in self.dicionario['localOferta']:
            self.nomeLocal.append(local['titulo'])

        for nivel in self.dicionario['nivelEnsino']:
            self.nomeNivel.append(nivel['titulo'])


        # Menu Lateral
        self.menu_frame = ctk.CTkFrame(self, width=500, corner_radius=0)
        self.menu_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.menu_frame.grid_rowconfigure(3, weight=1)
        self.menu_frame.grid_columnconfigure(0, weight=0)
        self.logo_label = ctk.CTkLabel(self.menu_frame, text="Rub Import", font=('Arial',20,"bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.tipo_label = ctk.CTkLabel(self.menu_frame, text="Tipo Importação:", anchor="w")
        self.tipo_label.grid(row=4, column=0, padx=0, pady=(5, 0))
        self.tipo_optionemenu = ctk.CTkComboBox(self.menu_frame, values=["Curso", "Oferta"], font=('Roboto',13), dropdown_fg_color=self.fg_selects, border_width=1, command=escolherImportacao)
        self.tipo_optionemenu.grid(row=5, column=0, padx=20, pady=(5, 0))

        self.scaling_label = ctk.CTkLabel(self.menu_frame, text="Escala:", anchor="w")
        self.scaling_label.grid(row=6, column=0, padx=0, pady=(5, 0))
        self.scaling_optionemenu = ctk.CTkComboBox(self.menu_frame, values=["80%", "90%", "100%", "110%", "120%"], command=self.mudarZoom, font=('Roboto',13), dropdown_fg_color=self.fg_selects, border_width=1)
        self.scaling_optionemenu.grid(row=7, column=0, padx=20, pady=(5, 15))

        self.label_oferta_pai = ctk.CTkLabel(self.menu_frame, text='Documentação', font=('Roboto',14))
        self.label_oferta_pai.grid(column=0, row=8, padx=(20,0), pady=(5,20), sticky='w')
        self.label_oferta = ctk.CTkButton(self.menu_frame, text='AQUI', command=abrirUrl, fg_color='transparent', hover='transparent', text_color='blue', width=20, font=('Roboto',14,'bold'))
        self.label_oferta.grid(column=0, row=8, padx=(115,0), pady=(5, 20), sticky='w')


        # Escolher arquivo
        self.arquivo = ctk.StringVar(master=self, value='Selecione o Arquivo')
        self.arquivo_input = ctk.CTkEntry(self, textvariable=self.arquivo, state='disable', font=('Roboto',13), border_color=self.border_color, border_width=self.borderInt)
        self.arquivo_input.grid(row=0, column=1, padx=(20, 0), pady=(40,0), sticky="new")
        
        img_upload = ctk.CTkImage(light_image=Image.open("./assets/cloud-arrow-up-solid-light.png"), dark_image=Image.open("./assets/cloud-arrow-up-solid-dark.png"), size=(20,15))
        self.btn_arquivo = ctk.CTkButton(master=self.arquivo_input, image=img_upload, text=None, command=abrirArquivo, font=('Roboto',14, 'bold'), width=25, fg_color='transparent', hover='transparent', text_color=self.text_color)
        self.btn_arquivo.grid(column=0, row=0, padx=2, pady=2, sticky='e')


        # Área importação
        self.switch_cod = ctk.CTkSwitch(master=self, text='Código Automático', onvalue="on", offvalue="off")
        self.switch_cod.grid(column=1, row=0, padx=(20,0), pady=(120,0), sticky='w')

        self.btn_import = ctk.CTkButton(self, text='Importar', font=('Roboto', 14), width=100, text_color='#ffffff',command=acionarImport, fg_color=self.btn_color, corner_radius=15)
        self.btn_import.grid(column=1, row=0, padx=0, pady=(120,0), sticky='e')


        # Terminal
        self.tab_terminal = ctk.CTkTabview(self, segmented_button_fg_color=self.color_unselected, segmented_button_unselected_color=self.color_unselected, segmented_button_selected_color=self.color_selected,segmented_button_unselected_hover_color=self.hover_butom, anchor='w', text_color=self.text_color, fg_color='transparent')
        self.tab_terminal.grid(row=2, column=1, padx=(20, 0), pady=0, sticky="ew")
        self.tab_terminal.add("Terminal")
        self.tab_terminal.tab("Terminal").grid_columnconfigure(0, weight=1)

        self.terminal = TerminalFrame(master=self.tab_terminal.tab("Terminal"), color=self.color_terminal)
        self.terminal.grid(row=0, column=0, padx=0, pady=0, sticky="nsew")


        # Tab Main
        self.tab_main = ctk.CTkTabview(self, height=215, width=300, segmented_button_fg_color=self.color_unselected, segmented_button_unselected_color=self.color_unselected, segmented_button_selected_color=self.color_selected,segmented_button_unselected_hover_color=self.hover_butom, anchor='w', text_color=self.text_color)
        self.tab_main.grid(row=0, column=2, padx=20, pady=(10, 0), sticky="nsew", columnspan=2, rowspan=2)
        self.tab_main.add("Configurações")
        self.tab_main.add("Geral")
        self.tab_main.tab("Configurações").grid_columnconfigure(0, weight=1) 
        self.tab_main.tab("Geral").grid_columnconfigure(0, weight=1)
        self.tab_main.tab("Geral").grid_rowconfigure(5, weight=1)

        # Tab main - Configurações
        self.comboMod = ctk.CTkComboBox(self.tab_main.tab('Configurações'), values=self.nomeMod, command=escolherModalidade, font=('Roboto',13), dropdown_fg_color=self.fg_selects, border_width=1, width=100)
        self.comboMod.grid(column=0, row=0, padx=(10), pady=(7,0), sticky='nsew')
        
        self.comboUni = ctk.CTkComboBox(self.tab_main.tab('Configurações'), values=self.nomeUnidade, command=escolherUnidade, font=('Roboto',13), dropdown_fg_color=self.fg_selects, border_width=1, width=100)
        self.comboUni.grid(column=0, row=1, padx=(10), pady=(7,0), sticky='nsew')
        
        self.comboNiv = ctk.CTkComboBox(self.tab_main.tab('Configurações'), values=self.nomeNivel, command=escolherNivel, font=('Roboto',13), dropdown_fg_color=self.fg_selects, border_width=1, width=100)
        self.comboNiv.grid(column=0, row=2, padx=(10), pady=(7,0), sticky='nsew')


        # Tab main - Geral
        self.btn_tranca = ctk.CTkButton(master=self.tab_main.tab("Geral"), image=img_lock, text=None, font=('Roboto',14, 'bold'), width=15, fg_color='transparent', hover='transparent', text_color=self.text_color, command=destrancar)
        self.btn_tranca.grid(column=1, row=0, padx=10, pady=0, sticky='e')

        self.linkStr = ctk.StringVar(master=self, value=self.link)
        self.link_input = ctk.CTkEntry(self.tab_main.tab("Geral"), textvariable=self.linkStr, state='disable', font=('Roboto',14), border_color=self.border_color, border_width=self.borderInt)
        self.link_input.grid(row=1, column=0, padx=10, pady=(2,0), sticky="nsew", columnspan=2)

        self.tokenStr = ctk.StringVar(master=self, value=self.token)
        self.token_input = ctk.CTkEntry(self.tab_main.tab("Geral"), textvariable=self.tokenStr, show="•", font=('Roboto',14), border_color=self.border_color, border_width=self.borderInt, state='disable')
        self.token_input.grid(row=2, column=0, padx=10, pady=(7,0), sticky="nsew", columnspan=2)

        self.btn_token = ctk.CTkButton(master=self.token_input, image=img_olho_oculto, text=None, font=('Roboto',14, 'bold'), height=10, width=10, fg_color='transparent', hover='transparent', command=desocultar)
        self.btn_token.grid(column=0, row=0, padx=2, pady=2, sticky='e')
        
        self.origemStr = ctk.StringVar(master=self, value=self.origem)
        self.origem_input = ctk.CTkEntry(self.tab_main.tab("Geral"), textvariable=self.origemStr, font=('Roboto',14), border_color=self.border_color, border_width=self.borderInt, state='disable')
        self.origem_input.grid(row=3, column=0, padx=10, pady=(5,0), sticky="nsew", columnspan=2)

        self.editar_btn = ctk.CTkButton(master=self.tab_main.tab("Geral"), text='Salvar', font=('Roboto',14), text_color='#ffffff', fg_color=self.btn_color, corner_radius=15, width=100, command=alterarChave, state='disable')

        self.barra_editar = ctk.CTkProgressBar(master=self.tab_main.tab("Geral"), height=3, orientation='horizontal', mode='inderteminate', progress_color=self.text_color, corner_radius=5)
        self.barra_editar.start()
        self.barra_editar.grid_forget()
        
        # Locais de Oferta
        self.switchLoc_oferta = LocalOfertaFrame(self, title='Selecione o Local de Oferta', values=self.nomeLocal)
        self.switchLoc_oferta.grid(row=2, column=2, padx=20, pady=0, sticky="ew", columnspan=2, rowspan=2)
        self.switchLoc_oferta.grid_columnconfigure(0, weight=1)
        

        # Valores Padrão
        self.scaling_optionemenu.set("100%")
        self.nomeTipoStr.set('Curso')

    def mudarZoom(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        ctk.set_widget_scaling(new_scaling_float)
    

class Conexao(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title('Rub Connection')
        position(win=self, width=375, height=135)
        self.columnconfigure(0, weight=1)
        self.iconbitmap('./assets/icon4.ico')

        def exit():
            self.destroy()

        if self._get_appearance_mode() == 'light':
            self.text_color = 'black'
            self.hover_butom = '#C0C0C0'
            self.border_color = 'black'
            self.btn_color = 'black'
            ctk.set_appearance_mode('light')
        else:
            self.text_color = '#ffffff'
            self.hover_butom = '#4F4F4F'
            self.border_color = 'silver'
            self.btn_color = 'blue'
            ctk.set_appearance_mode('dark')

        img_olho = ctk.CTkImage(light_image=Image.open("./assets/eye-solid-light.png"), dark_image=Image.open("./assets/eye-solid-dark.png"), size=(18,13))
        img_olho_oculto = ctk.CTkImage(light_image=Image.open("./assets/eye-slash-solid-light.png"), dark_image=Image.open("./assets/eye-slash-solid-dark.png"), size=(18,13))
        self.oculto = True

        def desocultar():
            if self.oculto == True:
                self.token_input.configure(show="")
                self.btn_token.configure(image=img_olho)
                self.oculto = False
            else:
                self.token_input.configure(show="•")
                self.btn_token.configure(image=img_olho_oculto)
                self.oculto = True
        
        def obter_resultados(result_queue):      
            try:
                permissao, url_valida, dicionario, dicionario_dois, link, origem, token = result_queue.get_nowait()
                dados = {'link':f'{link}','token':f'{token}','origem':origem}
                if permissao == False:
                    voltar_conexão()
                    threading.Thread(target=alerta_permissao())
                    logging.error(f"Token/Origem sem permissão para se conectar. Dados:{dados}")
                elif url_valida == False:
                    voltar_conexão()
                    threading.Thread(target=alerta_link())
                    logging.error(f"Não foi possível conectar, URL inválida. Dados:{dados}")
                else:
                    conexao.destroy()
                    window = App(dicionario=dicionario, dicionarioDois=dicionario_dois, link=link, origem=origem, token=token)
                    window.mainloop()
            except Exception as e:
                logging.error(f"Erro: {e}")

        def conectar():
            origem = int(self.origem_input.get())
            token = self.token_input.get()
            link_bruto = self.link_input.get()

            if link_bruto.count('/') == 2:
                link = link_bruto + '/'
            else:
                link = link_bruto
            

            self.update()
            self.link_input.grid_forget()
            self.token_input.grid_forget()
            self.btn_token.grid_forget()
            self.origem_input.grid_forget()
            self.cancelar_btn.grid_forget()
            self.connect_btn.grid_forget()
            self.update()
            self.Label = ctk.CTkLabel(self, text='Conectando...', font=('Roboto',14))
            self.Label.grid(column=0, row=0, sticky='nsew', padx=30, pady=(30,0))
            self.update()
            self.barLoad = ctk.CTkProgressBar(self, orientation='horizontal', mode='inderteminate', width=250, progress_color=self.btn_color)
            self.barLoad.start()
            self.update()
            self.barLoad.grid(column=0, row=1,sticky='nsew', padx=30, pady=(5,0))
            self.update()

            result_queue = Queue()
            
            threading.Thread(target=lambda:validarConexao(origem=origem, token=token, link=link, master=self, queue=result_queue)).start()
            self.after(15000, lambda: obter_resultados(result_queue))

        def checar_internet():
            url = 'https://www.google.com'
            timeout = 3
            try:
                rqt.get(url, timeout=timeout)
                return True
            except rqt.exceptions.ConnectionError:
                return False
        
        def alerta_conexao():
            msg = CTkMessagebox(title="Conexão Indisponíivel", message="Ops, algo está errado!\nNão foi possível carregar as informações.\nVerifique sua conexão e tente novamente.", icon="warning", font=('Roboto',14), justify='center', height=170, fg_color='transparent', fade_in_duration=450)

            if msg.get() == 'OK':
                self.quit()
            else:
                self.quit()

        
        def alerta_link():
            msg = CTkMessagebox(title="Link inválido", message="Ops, algo está errado!\nNão foi possível estabelecer uma conexçao.\nVerifique sua URL e tente novamente.", icon="cancel", font=('Roboto',14), justify='center', height=170, fg_color='transparent')
                    
        
        def voltar_conexão():
            self.barLoad.grid_forget()
            self.Label.grid_forget()
            self.update()
            self.link_input.grid(row=0, column=0, pady=(7,10), padx=7.5)
            self.token_input.grid(row=1, column=0, pady=(0,10), padx=7.5)
            self.btn_token.grid(column=0, row=0, padx=2, pady=2, sticky='e')
            self.origem_input.grid(row=2, column=0, pady=(0,10), padx=(0,200))
            self.connect_btn.grid(row=2, column=0, pady=(0,10), padx=(260,0))
            self.cancelar_btn.grid(row=2, column=0, pady=(0,10), padx=(65,0))
            self.update()

       
        self.conectado = checar_internet()
        if self.conectado == True:
            self.link_input = ctk.CTkEntry(master=self, placeholder_text='Link da base', font=('Roboto',14), width=350, border_color=self.border_color, border_width=1, placeholder_text_color=self.border_color)
            self.link_input.grid(row=0, column=0, pady=(13,10), padx=7.5)

            self.token_input = ctk.CTkEntry(master=self, placeholder_text='Token', font=('Roboto',14), show="•", width=350, border_color=self.border_color, border_width=1, placeholder_text_color=self.border_color)
            self.token_input.grid(row=1, column=0, pady=(0,10), padx=7.5)

            self.btn_token = ctk.CTkButton(master=self.token_input, image=img_olho_oculto, text=None, font=('Roboto',14, 'bold'), height=10, width=10, fg_color='transparent', hover='transparent', command=desocultar)
            self.btn_token.grid(column=0, row=0, padx=2, pady=2, sticky='e')

            self.origem_input = ctk.CTkEntry(master=self, placeholder_text='Origem', font=('Roboto',14), width=150, border_color=self.border_color, border_width=1, placeholder_text_color=self.border_color)
            self.origem_input.grid(row=2, column=0, pady=(0,10), padx=(0,200))

            self.connect_btn = ctk.CTkButton(master=self, text='Conectar', font=('Roboto',14, 'bold'), width=90, text_color='#ffffff', command=conectar, fg_color=self.btn_color)
            self.connect_btn.grid(row=2, column=0, pady=(0,10), padx=(260,0))

            self.cancelar_btn = ctk.CTkButton(master=self, text='Cancelar', font=('Roboto',14), width=90, border_width=1,fg_color='transparent', command=exit, hover_color=self.hover_butom, border_color=self.border_color, text_color=self.text_color)
            self.cancelar_btn.grid(row=2, column=0, pady=(0,10), padx=(65,0))

            self.resizable(width=False, height=False)
        else:
            alerta_conexao()
            logging.error(f"Não foi possível conectar, sem conexão com a internet.")


class LocalOfertaFrame(ctk.CTkScrollableFrame):
    def __init__(self, master, title, values):
        super().__init__(master, label_text=title, height=200)
        self.grid_columnconfigure(0, weight=1)
        self.values = values
        self.checkboxes = []
        self.checkboxPai = ctk.CTkSwitch(self, text='Selecionar Todos', command=self.marcaTodos, onvalue="on", offvalue="off", )
        self.checkboxPai.grid(row=0, column=0, padx=20, pady=(5, 0), sticky="w")
        for i, value in enumerate(self.values):
            checkbox = ctk.CTkSwitch(self, text=value, command=self.desmarcaPai)
            checkbox.grid(row=i+1, column=0, padx=20, pady=(5, 0), sticky="w")
            self.checkboxes.append(checkbox)
    def get(self):
        checked_checkboxes = []
        for checkbox in self.checkboxes:
            if checkbox.get() == 1:
                checked_checkboxes.append(checkbox.cget("text"))
        return checked_checkboxes
    
    def marcaTodos(self):
        if self.checkboxPai.get() == 'on':
            for checkbox in self.checkboxes:
                if checkbox.get() != 1:
                    checkbox.select(True)
        elif self.checkboxPai.get() == 'off':
            for checkbox in self.checkboxes:
                if checkbox.get() == 1:
                    checkbox.deselect()

    def desmarcaPai(self):
        if self.checkboxPai.get() == 'on':
            self.checkboxPai.deselect()
        todos = True
        for checkbox in self.checkboxes:
            if checkbox.get() != 1:
                todos = False
        
        if todos == True:
            self.checkboxPai.select(True)


class TerminalFrame(ctk.CTkScrollableFrame):
    def __init__(self, master, color):
        super().__init__(master, fg_color=color)
        self.grid_columnconfigure(0, weight=1)
        self.logs = []


def contador(arquivo):
    contador = 0
    for sheet in arquivo.sheetnames:
        for i, linha_arquivo in enumerate(arquivo[sheet].iter_rows(values_only=True)):
            if linha_arquivo[0] != None:
                contador = contador + 1
            else:
                break
        break
    return contador


def cadastrarCurso(file, link,  origem, token, tipo, dic, masterTerminal, master, dic_cod_curso):
    labelInit = ctk.CTkLabel(masterTerminal, text=f">>> Importação iniciada \n>>> Tipo: Curso", text_color='yellow', font=('Roboto Mono', 15))
    labelInit.grid(row=len(masterTerminal.logs), column=0, padx=5, pady=0, sticky='w')
    logging.info(f"Importação iniciada. Tipo: Curso")
    masterTerminal.logs.append(labelInit)
    master.update()
    arquivo = openpyxl.load_workbook(filename=f"{file}")
    quantidade = contador(arquivo)
    AdicionarContadorTerminal(quantidade, masterTerminal)
    master.update()
    sheet = arquivo.active
    index = len(dic)
    erro = 0
    for sheet_curso in arquivo.sheetnames:
        for i, linha_curso in enumerate(arquivo[sheet_curso].iter_rows(values_only=True)):
            l = i+1
            existe = False
            if linha_curso[0] != None:
                dados = {'codCurso':f'{linha_curso[1]}','nome':f'{linha_curso[0]}'}
                for e in dic: 
                    if tipo == 'off':
                        if linha_curso[1] in dic_cod_curso:
                            existe = True
                    if (tipo == 'off' and existe == False) or tipo == 'on':
                        if linha_curso[0] == e['nome'] and origem == e['origem']:
                            existe = True
            
                if existe == False:
                    if tipo == 'on':
                        index = index + 1
                    else:
                        index = linha_curso[1]
                    params = {
                        "nome":f"{linha_curso[0]}",
                        "origem":origem,
                        "token":f"{token}",
                        "codigo":f"{index}"
                    }

                    response = rqt.request('post', url=f"{link}api/Curso/cadastroCurso", json=params, timeout=3)
                    response.encoding
                    response.text
                    response_json = response.json()

                    if response_json['success'] == False:
                        sheet[f'D{l}'].value = 'Erro'
                        logging.error(f"Curso não cadastrada. Dados: {params} Resposta: {response_json}")
                        erro = erro + 1
                    else:
                        sheet[f'D{l}'].value = 'Cadastrado com Sucesso'
                        logging.info(f"[Success] Curso cadastrado com sucesso. Dados: {params} Resposta: {response_json}")
                else:
                    sheet[f'D{l}'].value = 'Curso já existe'
                    logging.error(f"Curso já existente no CRM. Dados: {dados}")
                    erro = erro + 1
            else:
                break
            MudarContadorTerminal(erro, l, quantidade, masterTerminal)
            master.update()
            arquivo.save(f'{file}')
            sleep(1.5)
        logging.info(f"Importação finalizada. Tipo: Curso - Sucesso: {quantidade-erro} Erro: {erro} Total: {quantidade}")
        break


def cadastrarOferta(file, link, origem, token, codModalidade, idModalidade, codUnidade, listLocal, idNivel, tipo, dic, masterTerminal, master, dic_curso):
        labelInit = ctk.CTkLabel(masterTerminal, text=f">>> Importação iniciada \n>>> Tipo: Oferta", text_color='yellow', font=('Roboto Mono', 15))
        labelInit.grid(row=len(masterTerminal.logs), column=0, padx=5, pady=0, sticky='w')
        logging.info(f"Importação iniciada. Tipo: Oferta")
        masterTerminal.logs.append(labelInit)
        master.update()
        arquivo = openpyxl.load_workbook(filename=f"{file}")
        quantidade = contador(arquivo)
        AdicionarContadorTerminal(quantidade, masterTerminal)
        master.update()
        sheet = arquivo.active
        index = len(dic)
        erro = 0
        for sheet_oferta in arquivo.sheetnames:
            for i, linha_oferta in enumerate(arquivo[sheet_oferta].iter_rows(values_only=True)):
                existe = False
                l = i+1
                if linha_oferta[0] != None:
                    dados = {'codCurso':f'{linha_oferta[0]}','nome':f'{linha_oferta[1]}','codOferta':f'{linha_oferta[2]}'}
                    for e in dic:
                        dic_local = []
                        for loc in e['locaisOferta']:
                            dic_local.append(loc['codigo'])
                        if linha_oferta[0] in dic_curso:
                            if tipo == 'off':
                                if linha_oferta[2] == e['codigo']:
                                    existe = True
                            if (tipo == 'off' and existe == False) or tipo == 'on':
                                if linha_oferta[1] == e['nome'] and idModalidade == int(e['modalidade']) and linha_oferta[0] == e['codCurso'] and idNivel == int(e['nivelEnsino']) and master.idUnidade == e['unidade'] and listLocal == dic_local and origem == int(e['origem']):
                                    existe = True  
                    if linha_oferta[0] in dic_curso:
                        if existe == False:
                            if tipo == 'on':
                                index = index + 1
                            else:
                                index = linha_oferta[2]
                            params = {
                                "codOferta": f"{index}",
                                "nome":f"{linha_oferta[1]}",
                                "modalidade":idModalidade,
                                "codModalidade":f"{codModalidade}",
                                "codCurso":f"{linha_oferta[0]}",
                                "codUnidade":f"{codUnidade}",
                                "codLocalOferta":listLocal,
                                "nivelEnsino":idNivel,
                                "origem":origem,
                                "token":f"{token}"
                            }
                            response = rqt.request('post', url=f"{link}api/Curso/cadastroOferta", json=params, timeout=3)
                            response.encoding
                            response.text
                            response_json = response.json()
                            if response_json['success'] == False:
                                sheet[f'D{l}'].value = 'Erro'
                                logging.error(f"Oferta não cadastrada. Dados: {params} Resposta: {response_json}")
                                erro = erro + 1
                            else:
                                sheet[f'D{l}'].value = 'Cadastrado com Sucesso'
                                logging.info(f"[Success] Oferta cadastrada com sucesso. Dados: {params} Resposta: {response_json}")
                        else:
                            sheet[f'D{l}'].value = 'Oferta já existe'
                            logging.error(f"Oferta já existente no CRM. Dados: {dados}")
                            erro = erro + 1
                    else:
                        sheet[f'D{l}'].value = 'Curso informado na oferta não existe'
                        logging.error(f"Curso informado na oferta não existe. Dados: {dados}")
                        erro = erro + 1
                else:
                    break
                MudarContadorTerminal(erro, l, quantidade, masterTerminal)
                master.update()
                arquivo.save(f'{file}')
                sleep(1.5)
            logging.info(f"Importação finalizada. Tipo: Oferta - Sucesso: {quantidade-erro} Erro: {erro} Total: {quantidade}")
            break


def AdicionarContadorTerminal(quantidade, master):
    index = len(master.logs)
    labelLog = ctk.CTkLabel(master, text=f">>> 0/{quantidade} Concluído", text_color='yellow', font=('Roboto Mono', 15))
    labelLog.grid(row=index, column=0, padx=5, pady=0, sticky='w')
    master.logs.append(labelLog)


def MudarContadorTerminal(erro, concluido, quant, master):
    label = master.logs[len(master.logs)-1]
    if concluido < quant:
        label.configure(text=f">>> {concluido}/{quant} Concluído")
    else:
        label.configure(text=f">>> {concluido}/{quant} Concluído\n>>> Sucesso: {quant-erro}", text_color='#00FF00')
        labelErro = ctk.CTkLabel(master, text=f">>> Erros: {erro}", text_color='red', font=('Roboto Mono', 15))
        labelErro.grid(row=len(master.logs), column=0, padx=5, pady=0, sticky='w')
        master.logs.append(labelErro)


def alerta_permissao():
    msg = CTkMessagebox(title="Sem permissão", message="Ops, algo está errado!\nToken e origem sem permissão", icon="cancel", font=('Roboto',14), justify='center', height=170, fg_color='transparent')


def validarConexao(origem, token, link, master, queue):

    dicionario = {}
    dicionario_dois = {}

    master.url_valida = True
    master.permissao = True

    params = {
        "origem":origem,
        "token":f"{token}"
    }

    try:
        response = rqt.request('get', url=f"{link}api/Registro/listarNiveisEnsino", json=params, timeout=1)
        response.encoding
        response.text
        listaNivel = response.json()
        sleep(1)
        response = rqt.request('get', url=f"{link}api/Curso/listarCursos", json=params, timeout=1)
        response.encoding
        response.text
        listaCurso = response.json()
        sleep(1)
        response = rqt.request('get', url=f"{link}api/Curso/listarOfertas", json=params, timeout=1)
        response.encoding
        response.text
        listaOferta = response.json()
        sleep(1)
        response = rqt.request('get', url=f"{link}api/Modalidade/listarModalidades", json=params, timeout=1)
        response.encoding
        response.text
        listaMod = response.json()
        sleep(1)
        response = rqt.request('get', url=f"{link}api/LocalOferta/listarLocaisOferta", json=params, timeout=1)
        response.encoding
        response.text
        listaLocal = response.json()
        sleep(1)
        response = rqt.request('get', url=f"{link}api/Unidade/listarUnidades", json=params, timeout=1)
        response.encoding
        response.text
        listaUnidade = response.json()
    except rqt.exceptions.ConnectionError:
        master.url_valida =  False

    if master.url_valida == False:
        queue.put((master.permissao,master.url_valida,dicionario, dicionario_dois, link, origem, token))
    elif listaNivel['success'] == False or listaLocal['success'] == False or listaMod['success'] == False or listaOferta['success'] == False or listaCurso['success'] == False or listaUnidade['success'] == False:
        master.permissao = False
        queue.put((master.permissao,master.url_valida,dicionario, dicionario_dois, link, origem, token))
    else:
        dicionario = {
            'modalidade': listaMod['dados'],
            'nivelEnsino': listaNivel['dados'],
            'unidade': listaUnidade['dados'],
            'localOferta': listaLocal['dados']
        }
        dicionario_dois = {
            'curso': listaCurso['dados'],
            'oferta': listaOferta['dados']
        }
        queue.put((master.permissao,master.url_valida,dicionario, dicionario_dois, link, origem, token))


if __name__ == "__main__":
    conexao = Conexao()
    conexao.mainloop()